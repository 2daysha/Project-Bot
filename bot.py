import telebot
from telebot import types
import openpyxl
import os
from dotenv import load_dotenv
from openpyxl.styles import PatternFill
load_dotenv()

TOKEN = os.getenv("TOKEN")
bot = telebot.TeleBot(TOKEN)

USER_STATE = {}

STUDENT_FILE = os.getenv("STUDENT_FILE")
TEACHER_FILE = os.getenv("TEACHER_FILE")
PROJECTS_FILE = os.getenv("PROJECTS_FILE")
PROPOSED_PROJECTS_FILE = os.getenv("PROPOSED_PROJECTS_FILE")
TEACHER_PASSWORD = os.getenv("TEACHER_PASSWORD")

def create_file_if_missing(file, headers):
    try:
        openpyxl.load_workbook(file)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file)


create_file_if_missing(STUDENT_FILE, ["ID", "Фамилия", "Имя", "Группа"])
create_file_if_missing(TEACHER_FILE, ["ID", "Фамилия", "Имя", "Отчество"])
create_file_if_missing(PROJECTS_FILE, ["ID Преподавателя", "Название проекта", "Фамилия студента",
                                       "Имя студента", "Группа студента", "Статус", "Прогресс","Оценка"])
create_file_if_missing(PROPOSED_PROJECTS_FILE, ["ID Студента", "Название проекта", "Фамилия студента",
                                                "Имя студента", "Группа студента", "ID Преподавателя"])

@bot.callback_query_handler(func=lambda call: call.data in ["register_student", "register_teacher"])
def register_user(call):
    user_id = call.from_user.id
    USER_STATE[user_id] = {'role': call.data.split("_")[1]}
    bot.send_message(user_id, "Введите вашу фамилию:")
    bot.register_next_step_handler(call.message, get_last_name)

def get_last_name(message):
    user_id = message.from_user.id
    USER_STATE[user_id]['last_name'] = message.text.strip()
    bot.send_message(user_id, "Введите ваше имя:")
    bot.register_next_step_handler(message, get_first_name)

def get_first_name(message):
    user_id = message.from_user.id
    USER_STATE[user_id]['first_name'] = message.text.strip()
    bot.send_message(user_id, "Введите ваше отчество (если есть):")
    bot.register_next_step_handler(message, get_middle_name)

def get_middle_name(message):
    user_id = message.from_user.id
    USER_STATE[user_id]['middle_name'] = message.text.strip()
    role = USER_STATE[user_id]['role']
    if role == "student":
        bot.send_message(user_id, "Введите вашу группу:")
        bot.register_next_step_handler(message, finalize_student_registration)
    elif role == "teacher":
        bot.send_message(user_id, "Введите пароль для регистрации преподавателем:")
        bot.register_next_step_handler(message, verify_teacher_password)

def get_teacher_subject(message):
    user_id = message.from_user.id
    USER_STATE[user_id]['subject'] = message.text.strip()
    finalize_teacher_registration(message)

def finalize_student_registration(message):
    user_id = message.from_user.id
    group = message.text.strip()

    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        ws.append([user_id, USER_STATE[user_id]['last_name'], USER_STATE[user_id]['first_name'], group])
        wb.save(STUDENT_FILE)
        bot.send_message(user_id, "Регистрация завершена! Вы зарегистрированы как студент.")
        send_student_menu(user_id)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при сохранении данных: {e}")
    finally:
        USER_STATE.pop(user_id, None)


def finalize_teacher_registration(message):
    user_id = message.from_user.id
    subject = USER_STATE[user_id]['subject']
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        ws.append([user_id, USER_STATE[user_id]['last_name'], USER_STATE[user_id]['first_name'], USER_STATE[user_id]['middle_name'], subject])
        wb.save(TEACHER_FILE)
        bot.send_message(user_id, "Регистрация завершена! Вы зарегистрированы как преподаватель.")
        send_teacher_menu(user_id)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при сохранении данных: {e}")
    finally:
        USER_STATE.pop(user_id, None)


def is_user_registered(user_id):
    return is_user_in_file(user_id, STUDENT_FILE) or is_user_in_file(user_id, TEACHER_FILE)


def is_teacher(user_id):
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return True
        return False
    except Exception as e:
        print(f"Ошибка проверки роли преподавателя: {e}")
        return False


def is_student(user_id):
    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return True
        return False
    except Exception as e:
        print(f"Ошибка проверки роли студента: {e}")
        return False


def is_user_in_file(user_id, file):
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(user_id):
                return True
        return False
    except FileNotFoundError:
        return False


@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.from_user.id
    if is_user_registered(user_id):
        send_main_menu(user_id)
    else:
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Зарегистрироваться как студент", callback_data="register_student"))
        markup.add(types.InlineKeyboardButton("Зарегистрироваться как преподаватель", callback_data="register_teacher"))
        bot.send_message(user_id, "Привет! Я помогу тебе организовать работу с проектами. Выберите роль:",
                         reply_markup=markup)
                     
def send_main_menu(user_id):
    role = get_user_role(user_id)
    if role == 'student':
        send_student_menu(user_id)
    elif role == 'teacher':
        send_teacher_menu(user_id)


def get_user_role(user_id):
    if is_user_in_file(user_id, STUDENT_FILE):
        return 'student'
    elif is_user_in_file(user_id, TEACHER_FILE):
        return 'teacher'
    return None


@bot.callback_query_handler(func=lambda call: call.data == "student_menu")
def send_student_menu(user_id):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Связаться с преподавателем", callback_data="contact_teacher"))
    markup.add(types.InlineKeyboardButton("Предложить тему проекта", callback_data="suggest_project"))
    markup.add(types.InlineKeyboardButton("Мои проекты", callback_data="my_projects"))
    markup.add(types.InlineKeyboardButton("Вернуться в меню", callback_data="back_to_main_menu"))
    bot.send_message(user_id, "Добро пожаловать, студент! Выберите действие:", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data == "teacher_menu")
def send_teacher_menu(user_id):
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Добавить проект", callback_data="add_project"))
    markup.add(types.InlineKeyboardButton("Поиск проекта", callback_data="search_project"))
    markup.add(types.InlineKeyboardButton("Скачать отчет", callback_data="download_report"))
    markup.add(types.InlineKeyboardButton("Изменить статус проекта", callback_data="change_status"))
    markup.add(types.InlineKeyboardButton("Оценить проект", callback_data="evaluate_project"))
    markup.add(types.InlineKeyboardButton("Отметить прогресс", callback_data="mark_progress"))
    markup.add(types.InlineKeyboardButton("Вернуться в меню", callback_data="back_to_main_menu"))
    bot.send_message(user_id, "Добро пожаловать, преподаватель! Выберите действие:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "back_to_main_menu")
def back_to_main_menu(call):
    user_id = call.from_user.id
    send_main_menu(user_id)


def initialize_projects_file():
    if not os.path.exists(PROJECTS_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects"
        ws.append(["Project ID", "Title", "Description", "Teacher ID", "Student ID", "Status", "Progress", "Evaluation"])
        wb.save(PROJECTS_FILE)



def get_student_data(student_id):
    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(student_id):
                return row[1], row[2], row[3]
        return None
    except Exception as e:
        print(f"Ошибка при получении данных студента: {e}")
        return None

def get_teacher_name(teacher_id):
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(teacher_id):
                return f"{row[1]} {row[2]}"
        return "Неизвестно"
    except Exception as e:
        print(f"Ошибка при получении ФИО преподавателя: {e}")
        return "Неизвестно"


@bot.callback_query_handler(func=lambda call: call.data == "add_project")
def add_project_handler(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return

    bot.send_message(user_id, "Введите название проекта:")
    bot.register_next_step_handler(call.message, get_project_title)


def get_project_title(message):
    user_id = message.from_user.id
    project_title = message.text.strip()

    bot.send_message(user_id, "Введите описание проекта:")
    bot.register_next_step_handler(message, get_project_description, project_title)


def get_project_description(message, project_title):
    user_id = message.from_user.id
    project_description = message.text.strip()

    bot.send_message(user_id, "Выберите преподавателя, который будет руководить проектом")

    try:
        markup = types.InlineKeyboardMarkup()
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            teacher_id, last_name, first_name, *_ = row
            teacher_name = f"{last_name} {first_name}"
            markup.add(types.InlineKeyboardButton(teacher_name, callback_data=f"teacher_{teacher_id}"))
        bot.send_message(user_id, "Список преподавателей:", reply_markup=markup)
        USER_STATE[user_id] = {"project_title": project_title, "project_description": project_description}
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при загрузке списка преподавателей: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("teacher_"))
def select_teacher_for_project(call):
    user_id = call.from_user.id
    teacher_id = call.data.split("_")[1]

    if user_id not in USER_STATE:
        bot.send_message(user_id, "Произошла ошибка. Начните добавление проекта заново.")
        return

    USER_STATE[user_id]["teacher_id"] = teacher_id
    bot.send_message(user_id, "Введите имя и фамилию первого студента (например, Иван Иванов):")
    bot.register_next_step_handler(call.message, get_student_by_name)


def get_student_by_name(message):
    user_id = message.from_user.id
    student_name = message.text.strip()

    if user_id not in USER_STATE:
        bot.send_message(user_id, "Произошла ошибка. Начните добавление проекта заново.")
        return

    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        student_data = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if f"{row[1]} {row[2]}".lower() == student_name.lower():
                student_data = row
                break

        if student_data:
            student_lastname, student_firstname, student_group = student_data[1], student_data[2], student_data[3]
            if "students" not in USER_STATE[user_id]:
                USER_STATE[user_id]["students"] = []
                USER_STATE[user_id]["students"].append((student_lastname, student_firstname, student_group))
                bot.send_message(user_id, f"Добавлен студент: {student_lastname} {student_firstname}, группа: {student_group}.")
                bot.send_message(user_id, "Введите имя и фамилию следующего студента или 'стоп' для завершения:")
                bot.register_next_step_handler(message, add_next_student)
        else:
            bot.send_message(user_id, "Студент с таким именем не найден. Попробуйте снова.")
            bot.register_next_step_handler(message, get_student_by_name)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске студента: {e}")


def add_next_student(message):
    user_id = message.from_user.id
    student_name = message.text.strip()

    if student_name.lower() == "стоп":
        bot.send_message(user_id, "Введите статус проекта (например, 'Создан'):")
        bot.register_next_step_handler(message, finalize_project_addition)
        return

    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        student_data = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if f"{row[1]} {row[2]}".lower() == student_name.lower():
                student_data = row
                break

        if student_data:
            student_lastname, student_firstname, student_group = student_data[1], student_data[2], student_data[3]
            USER_STATE[user_id]["students"].append((student_lastname, student_firstname, student_group))
            bot.send_message(user_id, f"Добавлен студент: {student_lastname} {student_firstname}, группа: {student_group}.")
        else:
            bot.send_message(user_id, "Студент с таким именем не найден. Попробуйте снова.")

        bot.send_message(user_id, "Введите имя и фамилию следующего студента или 'стоп' для завершения:")
        bot.register_next_step_handler(message, add_next_student)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при добавлении студента: {e}")


def finalize_project_addition(message):
    user_id = message.from_user.id
    project_status = message.text.strip()

    try:
        data = USER_STATE[user_id]
        teacher_id = data["teacher_id"]
        teacher_name = get_teacher_name(teacher_id)
        project_title = data["project_title"]
        project_description = data["project_description"]
        students = data["students"]

        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active

        if ws.max_row == 1:
            ws.append(["ФИО Преподавателя", "Название проекта", "Фамилия студента", "Имя студента", "Группа студента",
                       "Статус", "Прогресс", "Оценка"])

        for student in students:
            student_lastname, student_firstname, student_group = student
            ws.append([teacher_name, project_title, student_lastname, student_firstname, student_group, project_status, "", "Не оценено"])

        wb.save(PROJECTS_FILE)
        bot.send_message(user_id, f"Проект '{project_title}' успешно добавлен.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при добавлении проекта: {e}")
    finally:
        USER_STATE.pop(user_id, None)

@bot.callback_query_handler(func=lambda call: call.data == "search_project")
def search_project_handler(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return
    else:
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Поиск по названию", callback_data="search_by_title"))
        markup.add(types.InlineKeyboardButton("Поиск по статусу проекта", callback_data="search_by_status"))
        bot.send_message(user_id, "Выберите способ поиска проекта:", reply_markup=markup)


def search_project(message):
    user_id = message.from_user.id
    search_query = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        found_projects = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) != 8:
                continue

            project_id, project_name, student_lastname, student_firstname, student_group, project_status, progress, evaluation = row

            if (str(search_query) in str(project_id)) or (search_query.lower() in str(project_name).lower()) or (search_query.lower() in str(project_status).lower()):
                participants = found_projects.get(project_id, [])
                participants.append(f"{student_lastname} {student_firstname} ({student_group})")
                found_projects[project_id] = participants

        if found_projects:
            result_message = "Найденные проекты:\n"
            for project_id, participants in found_projects.items():
                wb = openpyxl.load_workbook(PROJECTS_FILE)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == project_id:
                        project_name = row[1]
                        project_status = row[5]
                        progress = row[6]
                        evaluation = row[7]
                        break

                result_message += f"Название: {project_name}, Участники: {', '.join(set(participants))}, Статус: {project_status}, Прогресс: {progress}, Оценка: {evaluation}\n"

            bot.send_message(user_id, result_message)
        else:
            bot.send_message(user_id, "Проекты по вашему запросу не найдены.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проектов: {e}")




def get_student_group(user_id):
    try:
        wb = openpyxl.load_workbook("students.xlsx")
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return row[2]

        return "Группа не указана"
    except Exception as e:
        print(f"Ошибка при загрузке группы: {e}")
        return "Группа не указана"


@bot.callback_query_handler(func=lambda call: call.data == "my_projects")
def my_projects(call):
    user_id = call.from_user.id
    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        student_projects = []

        student_lastname = None
        student_firstname = None
        student_data = get_student_data(user_id)
        if not student_data:
            bot.send_message(user_id, "Не удалось найти вашу информацию в базе данных.")
            return

        student_lastname, student_firstname, student_group = student_data

        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) != 8:
                continue
            if row[2].strip().lower() == student_lastname.lower() and row[3].strip().lower() == student_firstname.lower():
                project_info = (
                    f"Название: {row[1]}, \n"
                    f"Преподаватель: {row[0]}, \n"
                    f"Группа студента: {row[4]}, \n"
                    f"Статус: {row[5]}, \n"
                    f"Прогресс: {row[6]}, \n"
                    f"Оценка: {row[7]}\n"
                )
                student_projects.append(project_info)

        if student_projects:
            bot.send_message(user_id, "Ваши проекты:\n\n" + "\n\n".join(student_projects))
        else:
            bot.send_message(user_id, "У вас нет проектов.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при загрузке проектов: {e}")


def get_student_data(user_id):
    try:
        wb = openpyxl.load_workbook(STUDENT_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                return row[1], row[2], row[3]  # Фамилия, Имя, Группа
        return None
    except Exception as e:
        print(f"Ошибка при получении данных студента: {e}")
        return None


@bot.callback_query_handler(func=lambda call: call.data == "change_status")
def change_status(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return
    bot.send_message(user_id, "Введите название проекта, для которого хотите изменить статус:")
    bot.register_next_step_handler(call.message, get_project_by_title_for_status)

def get_project_by_title_for_status(message):
    user_id = message.from_user.id
    project_title = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        projects_found = []

        # Ищем проекты по названию
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1].lower() == project_title.lower():  # Название проекта вторая колонка
                projects_found.append(row)

        if not projects_found:
            bot.send_message(user_id, "Проект с таким названием не найден. Попробуйте снова.")
            return

        if len(projects_found) == 1:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("Создан", callback_data=f"status_created_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("В процессе", callback_data=f"status_in_progress_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("Завершен", callback_data=f"status_completed_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("Отменен", callback_data=f"status_cancelled_{projects_found[0][1]}"))
            bot.send_message(user_id, "Выберите новый статус проекта:", reply_markup=markup)
        else:
            markup = types.InlineKeyboardMarkup()
            for idx, project in enumerate(projects_found):
                student_info = f"{project[2]} {project[3]} ({project[4]})"
                markup.add(types.InlineKeyboardButton(f"{project[1]} - {student_info}", callback_data=f"choose_{idx}"))
            bot.send_message(user_id, "Найдено несколько проектов с таким названием. Выберите проект:", reply_markup=markup)
            USER_STATE[user_id] = {'projects_found': projects_found}
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проекта: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("status_"))
def set_project_status(call):
    user_id = call.from_user.id
    status_action, project_title = call.data.split("_")[1:3]
    status_map = {
        "created": "Создан",
        "in_progress": "В процессе",
        "completed": "Завершен",
        "cancelled": "Отменен",
    }
    new_status = status_map.get(status_action, "Неизвестно")

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value.lower() == project_title.lower():
                row[5].value = new_status
                wb.save(PROJECTS_FILE)
                bot.send_message(user_id, f"Статус проекта '{project_title}' успешно обновлён на '{new_status}'.")
                return
        bot.send_message(user_id, "Проект не найден. Попробуйте снова.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при обновлении статуса проекта: {e}")




@bot.callback_query_handler(func=lambda call: call.data == "evaluate_project")
def evaluate_project(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return
    bot.send_message(user_id, "Введите название проекта, который хотите оценить:")
    bot.register_next_step_handler(call.message, get_project_by_title_for_evaluation)


def get_project_by_title_for_evaluation(message):
    user_id = message.from_user.id
    project_title = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        projects_found = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1].lower() == project_title.lower():
                projects_found.append(row)

        if not projects_found:
            bot.send_message(user_id, "Проект с таким названием не найден. Попробуйте снова.")
            return

        if len(projects_found) == 1:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("Отлично", callback_data=f"evaluation_excellent_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("Хорошо", callback_data=f"evaluation_good_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("Удовлетворительно", callback_data=f"evaluation_satisfactory_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("Неуд", callback_data=f"evaluation_unsatisfactory_{projects_found[0][1]}"))
            bot.send_message(user_id, "Выберите оценку для проекта:", reply_markup=markup)
        else:
            markup = types.InlineKeyboardMarkup()
            for idx, project in enumerate(projects_found):
                student_info = f"{project[2]} {project[3]} ({project[4]})"
                markup.add(types.InlineKeyboardButton(f"{project[1]} - {student_info}", callback_data=f"choose_eval_{idx}"))
            bot.send_message(user_id, "Найдено несколько проектов с таким названием. Выберите проект:", reply_markup=markup)
            USER_STATE[user_id] = {'projects_found': projects_found}
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проекта: {e}")



@bot.callback_query_handler(func=lambda call: call.data.startswith("choose_eval_"))
def choose_project_for_evaluation(call):
    user_id = call.from_user.id
    if user_id not in USER_STATE or 'projects_found' not in USER_STATE[user_id]:
        bot.send_message(user_id, "Ошибка! Попробуйте снова.")
        return

    project_index = int(call.data.split("_")[2])
    project = USER_STATE[user_id]['projects_found'][project_index]
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Отлично", callback_data=f"evaluation_excellent_{project[1]}"))
    markup.add(types.InlineKeyboardButton("Хорошо", callback_data=f"evaluation_good_{project[1]}"))
    markup.add(types.InlineKeyboardButton("Удовлетворительно", callback_data=f"evaluation_satisfactory_{project[1]}"))
    markup.add(types.InlineKeyboardButton("Неуд", callback_data=f"evaluation_unsatisfactory_{project[1]}"))
    bot.send_message(user_id, "Выберите оценку для проекта:", reply_markup=markup)
    USER_STATE.pop(user_id, None)

from openpyxl.styles import PatternFill


@bot.callback_query_handler(func=lambda call: call.data == "evaluate_project")
def evaluate_project_handler(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return
    bot.send_message(user_id, "Введите название проекта, который хотите оценить:")
    bot.register_next_step_handler(call.message, get_project_by_title_for_evaluation)

def get_project_by_title_for_evaluation(message):
    user_id = message.from_user.id
    project_title = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        projects_found = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1].lower() == project_title.lower():
                projects_found.append(row)

        if not projects_found:
            bot.send_message(user_id, "Проект с таким названием не найден. Попробуйте снова.")
            return

        if len(projects_found) == 1:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("Отлично", callback_data=f"evaluate_excellent_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("Хорошо", callback_data=f"evaluate_good_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("Удовлетворительно", callback_data=f"evaluate_satisfactory_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("Неуд", callback_data=f"evaluate_poor_{projects_found[0][1]}"))
            bot.send_message(user_id, "Выберите оценку для проекта:", reply_markup=markup)
        else:
            markup = types.InlineKeyboardMarkup()
            for idx, project in enumerate(projects_found):
                student_info = f"{project[2]} {project[3]} ({project[4]})"
                markup.add(types.InlineKeyboardButton(f"{project[1]} - {student_info}", callback_data=f"choose_eval_{idx}"))
            bot.send_message(user_id, "Найдено несколько проектов с таким названием. Выберите проект:", reply_markup=markup)
            USER_STATE[user_id] = {'projects_found': projects_found}
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проекта: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("choose_eval_"))
def choose_project_for_evaluation(call):
    user_id = call.from_user.id
    if user_id not in USER_STATE or 'projects_found' not in USER_STATE[user_id]:
        bot.send_message(user_id, "Ошибка! Попробуйте снова.")
        return

    project_index = int(call.data.split("_")[2])
    project = USER_STATE[user_id]['projects_found'][project_index]
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Отлично", callback_data=f"evaluate_excellent_{project[1]}"))
    markup.add(types.InlineKeyboardButton("Хорошо", callback_data=f"evaluate_good_{project[1]}"))
    markup.add(types.InlineKeyboardButton("Удовлетворительно", callback_data=f"evaluate_satisfactory_{project[1]}"))
    markup.add(types.InlineKeyboardButton("Неуд", callback_data=f"evaluate_poor_{project[1]}"))
    bot.send_message(user_id, "Выберите оценку для проекта:", reply_markup=markup)
    USER_STATE.pop(user_id, None)

@bot.callback_query_handler(func=lambda call: call.data.startswith("evaluate_"))
def set_project_evaluation(call):
    user_id = call.from_user.id
    evaluation_action, project_title = call.data.split("_")[1:3]
    evaluation_map = {
        "excellent": ("Отлично", "90EE90"),
        "good": ("Хорошо", "00BFFF"),
        "satisfactory": ("Удовлетворительно", "FFA500"),
        "poor": ("Неуд", "FF0000")
    }
    new_evaluation, color_code = evaluation_map.get(evaluation_action, ("Неизвестно", "FFFFFF"))

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value.lower() == project_title.lower():
                row[7].value = new_evaluation
                fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                row[7].fill = fill
                wb.save(PROJECTS_FILE)
                bot.send_message(user_id, f"Оценка проекта '{project_title}' успешно обновлена на '{new_evaluation}'.")
                return
        bot.send_message(user_id, "Проект не найден. Попробуйте снова.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при обновлении оценки проекта: {e}")

@bot.callback_query_handler(func=lambda call: call.data == "download_report")
def download_report(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return

    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Промежуточный отчёт", callback_data="interim_report"))
    markup.add(types.InlineKeyboardButton("Итоговый отчёт", callback_data="final_report"))
    bot.send_message(user_id, "Выберите тип отчёта, который вы хотите скачать:", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "suggest_project")
def suggest_project(call):
    user_id = call.from_user.id
    USER_STATE[user_id] = {"action": "suggest_project"}
    bot.send_message(user_id, "Выберите преподавателя, которому хотите предложить тему проекта:")

    markup = types.InlineKeyboardMarkup()
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            teacher_id = row[0]
            teacher_name = f"{row[1]} {row[2]}"
            markup.add(types.InlineKeyboardButton(teacher_name, callback_data=f"suggest_teacher_{teacher_id}"))
        bot.send_message(user_id, "Выберите преподавателя из списка:", reply_markup=markup)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при загрузке преподавателей: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("suggest_teacher_"))
def teacher_selected_for_suggestion(call):
    teacher_id = call.data.split("_")[2]
    user_id = call.from_user.id

    if user_id not in USER_STATE or USER_STATE[user_id].get("action") != "suggest_project":
        bot.send_message(user_id, "Произошла ошибка. Начните процесс заново.")
        return

    USER_STATE[user_id]["teacher_id"] = teacher_id
    bot.send_message(user_id, "Введите тему проекта, которую вы хотите предложить:")
    bot.register_next_step_handler(call.message, handle_project_suggestion)


def handle_project_suggestion(message):
    user_id = message.from_user.id
    suggestion = message.text.strip()

    if user_id not in USER_STATE or "teacher_id" not in USER_STATE[user_id]:
        bot.send_message(user_id, "Произошла ошибка. Начните процесс заново.")
        return

    teacher_id = USER_STATE[user_id]["teacher_id"]

    try:
        wb_students = openpyxl.load_workbook(STUDENT_FILE)
        ws_students = wb_students.active
        student_info = None
        for row in ws_students.iter_rows(min_row=2, values_only=True):
            if row[0] == user_id:
                student_info = row
                break

        if not student_info:
            bot.send_message(user_id, "Ошибка: ваша регистрация как студента не найдена.")
            return

        student_lastname, student_firstname, student_group = student_info[1:4]

        wb_proposed = openpyxl.load_workbook(PROPOSED_PROJECTS_FILE)
        ws_proposed = wb_proposed.active
        ws_proposed.append([user_id, suggestion, student_lastname, student_firstname, student_group, teacher_id])
        wb_proposed.save(PROPOSED_PROJECTS_FILE)

        bot.send_message(user_id, f"Тема проекта '{suggestion}' успешно предложена преподавателю!")

        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Одобрить", callback_data=f"suggest_approve_{user_id}"))
        markup.add(types.InlineKeyboardButton("Не одобрить", callback_data=f"suggest_reject_{user_id}"))
        bot.send_message(
            teacher_id,
            f"Студент {student_lastname} {student_firstname}, группа {student_group} предложил тему: '{suggestion}'.\n"
            "Нажмите одну из кнопок ниже для одобрения или отклонения этой темы.",
            reply_markup=markup
        )
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при предложении темы: {e}")
    finally:
        USER_STATE.pop(user_id, None)


@bot.callback_query_handler(func=lambda call: call.data.startswith("suggest_approve_"))
def approve_project_suggestion(call):
    student_id = call.data.split("_")[2]
    teacher_id = call.from_user.id

    try:
        wb_proposed = openpyxl.load_workbook(PROPOSED_PROJECTS_FILE)
        ws_proposed = wb_proposed.active

        for row in ws_proposed.iter_rows(min_row=2, values_only=False):
            if str(row[0].value) == student_id:
                project_title = row[1].value
                student_lastname = row[2].value
                student_firstname = row[3].value
                student_group = row[4].value

                wb_projects = openpyxl.load_workbook(PROJECTS_FILE)
                ws_projects = wb_projects.active

                if ws_projects.max_row == 1:
                    ws_projects.append(["ФИО Преподавателя", "Название проекта", "Фамилия студента",
                                        "Имя студента", "Группа студента", "Статус", "Прогресс", "Оценка"])

                ws_projects.append([
                    get_teacher_name(teacher_id), project_title, student_lastname, student_firstname, student_group,
                    "Создан", "", "Не оценено"
                ])
                wb_projects.save(PROJECTS_FILE)

                ws_proposed.delete_rows(row[0].row)
                wb_proposed.save(PROPOSED_PROJECTS_FILE)

                bot.send_message(student_id, f"Ваш проект '{project_title}' был одобрен преподавателем!")
                bot.send_message(teacher_id, f"Проект '{project_title}' успешно одобрен.")
                return

        bot.send_message(teacher_id, "Проект с таким студентом не найден. Попробуйте снова.")
    except Exception as e:
        bot.send_message(teacher_id, f"Ошибка при одобрении проекта: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("suggest_reject_"))
def reject_project_suggestion(call):
    student_id = call.data.split("_")[2]
    teacher_id = call.from_user.id

    bot.send_message(teacher_id, "Введите комментарий для отклонения темы проекта:")
    USER_STATE[teacher_id] = {"action": "reject_suggestion", "student_id": student_id}
    bot.register_next_step_handler(call.message, handle_rejection_comment)


def handle_rejection_comment(message):
    teacher_id = message.from_user.id
    comment = message.text.strip()

    if teacher_id not in USER_STATE or USER_STATE[teacher_id].get("action") != "reject_suggestion":
        bot.send_message(teacher_id, "Произошла ошибка. Попробуйте снова.")
        return

    student_id = USER_STATE[teacher_id]["student_id"]

    try:
        wb_proposed = openpyxl.load_workbook(PROPOSED_PROJECTS_FILE)
        ws_proposed = wb_proposed.active

        for row in ws_proposed.iter_rows(min_row=2, values_only=False):
            if str(row[0].value) == student_id:

                project_title = row[1].value
                bot.send_message(student_id, f"Ваш проект '{project_title}' был отклонён преподавателем.\nКомментарий: {comment}")

                ws_proposed.delete_rows(row[0].row)
                wb_proposed.save(PROPOSED_PROJECTS_FILE)

                bot.send_message(teacher_id, f"Проект '{project_title}' успешно отклонён с комментарием.")
                return

        bot.send_message(teacher_id, "Проект с таким студентом не найден. Попробуйте снова.")
    except Exception as e:
        bot.send_message(teacher_id, f"Ошибка при отклонении проекта: {e}")
    finally:
        USER_STATE.pop(teacher_id, None)
    
""" ---------------------------"""

@bot.callback_query_handler(func=lambda call: call.data == "mark_progress")
def mark_progress_handler(call):
    user_id = call.from_user.id
    if not is_teacher(user_id):
        bot.send_message(user_id, "Ошибка! Вы не преподаватель.")
        return
    bot.send_message(user_id, "Введите название проекта, для которого хотите отметить прогресс:")
    bot.register_next_step_handler(call.message, get_project_by_title_for_progress)

def get_project_by_title_for_progress(message):
    user_id = message.from_user.id
    project_title = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        projects_found = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1].lower() == project_title.lower():
                projects_found.append(row)

        if not projects_found:
            bot.send_message(user_id, "Проект с таким названием не найден. Попробуйте снова.")
            return

        if len(projects_found) == 1:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("➕ ", callback_data=f"progress_plus_{projects_found[0][1]}"))
            markup.add(types.InlineKeyboardButton("➖", callback_data=f"progress_minus_{projects_found[0][1]}"))
            bot.send_message(user_id, "Выберите действие для отметки прогресса:", reply_markup=markup)
        else:
            markup = types.InlineKeyboardMarkup()
            for idx, project in enumerate(projects_found):
                student_info = f"{project[2]} {project[3]} ({project[4]})"
                markup.add(types.InlineKeyboardButton(f"{project[1]} - {student_info}", callback_data=f"choose_prog_{idx}"))
            bot.send_message(user_id, "Найдено несколько проектов с таким названием. Выберите проект:", reply_markup=markup)
            USER_STATE[user_id] = {'projects_found': projects_found}
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проекта: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("choose_prog_"))
def choose_project_for_progress(call):
    user_id = call.from_user.id
    if user_id not in USER_STATE or 'projects_found' not in USER_STATE[user_id]:
        bot.send_message(user_id, "Ошибка! Попробуйте снова.")
        return 
    project_index = int(call.data.split("_")[2])
    project = USER_STATE[user_id]['projects_found'][project_index]
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("+", callback_data=f"progress_plus_{project[1]}"))
    markup.add(types.InlineKeyboardButton("-", callback_data=f"progress_minus_{project[1]}"))
    bot.send_message(user_id, "Выберите действие для отметки прогресса:", reply_markup=markup)
    USER_STATE.pop(user_id, None)

@bot.callback_query_handler(func=lambda call: call.data.startswith("progress_"))
def set_project_progress(call):
    user_id = call.from_user.id
    action, project_title = call.data.split("_")[1:3]
    progress_change = "+" if action == "plus" else "-"

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value.lower() == project_title.lower():
                current_progress = row[6].value
                if current_progress is None:
                    current_progress = ""
                row[6].value = current_progress + progress_change
                wb.save(PROJECTS_FILE)
                bot.send_message(user_id, f"Прогресс проекта '{project_title}' успешно обновлён на '{current_progress + progress_change}'.")
                return
        bot.send_message(user_id, "Проект не найден. Попробуйте снова.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при обновлении прогресса проекта: {e}")

@bot.callback_query_handler(func=lambda call: call.data.startswith("choose_"))
def choose_project_for_status(call):
    user_id = call.from_user.id
    if user_id not in USER_STATE or 'projects_found' not in USER_STATE[user_id]:
        bot.send_message(user_id, "Ошибка! Попробуйте снова.")
        return

    project_index = int(call.data.split("_")[1])
    project = USER_STATE[user_id]['projects_found'][project_index]
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Создан", callback_data=f"status_created_{project[1]}"))
    markup.add(types.InlineKeyboardButton("Одобрен", callback_data=f"status_approved_{project[1]}"))
    markup.add(types.InlineKeyboardButton("Завершен", callback_data=f"status_completed_{project[1]}"))
    markup.add(types.InlineKeyboardButton("Отменен", callback_data=f"status_cancelled_{project[1]}"))
    bot.send_message(user_id, "Выберите новый статус проекта:", reply_markup=markup)
    USER_STATE.pop(user_id, None)


def verify_teacher_password(message):
    user_id = message.from_user.id
    password = message.text.strip()
    if password == TEACHER_PASSWORD:
        bot.send_message(user_id, "Пароль принят. Введите ваш предмет:")
        bot.register_next_step_handler(message, get_teacher_subject)
    else:
        bot.send_message(user_id, "В доступе отказано.")
        USER_STATE.pop(user_id, None)

def get_teacher_subject(message):
    user_id = message.from_user.id
    USER_STATE[user_id]['subject'] = message.text.strip()
    finalize_teacher_registration(message)

@bot.callback_query_handler(func=lambda call: call.data == "search_by_title")
def search_by_title_handler(call):
    user_id = call.from_user.id
    bot.send_message(user_id, "Введите название проекта для поиска:")
    bot.register_next_step_handler(call.message, search_project_by_title)
    
def search_project_by_title(message):
    user_id = message.from_user.id
    project_title = message.text.strip()

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        projects_found = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if project_title.lower() in row[1].lower():
                projects_found.append(row)

        if projects_found:
            result_message = "Найденные проекты:\n"
            for project in projects_found:
                result_message += (
                    f"Название: {project[1]}\n"
                    f"Преподаватель: {project[0]}\n"
                    f"Статус: {project[5]}\n"
                    f"Прогресс: {project[6]}\n"
                    f"Оценка: {project[7]}\n\n"
                )
            bot.send_message(user_id, result_message)
        else:
            bot.send_message(user_id, "Проектов с таким названием не найдено.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проектов: {e}")

@bot.callback_query_handler(func=lambda call: call.data == "search_by_status")
def search_by_status_handler(call):
    user_id = call.from_user.id
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Создан", callback_data="search_status_created"))
    markup.add(types.InlineKeyboardButton("Одобрен", callback_data="search_status_approved"))
    markup.add(types.InlineKeyboardButton("Завершен", callback_data="search_status_completed"))
    markup.add(types.InlineKeyboardButton("Отменен", callback_data="search_status_cancelled"))
    bot.send_message(user_id, "Выберите статус проекта для поиска:", reply_markup=markup)

@bot.callback_query_handler(func=lambda call: call.data.startswith("search_status_"))
def search_project_by_status(call):
    user_id = call.from_user.id
    status_map = { "created": "Создан", "approved": "Одобрен", "completed": "Завершен", "cancelled": "Отменен" }
    status_action = call.data.split("_")[2]
    project_status = status_map.get(status_action, "")

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active
        projects_found = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if project_status.lower() in row[5].lower():
                projects_found.append(row)

        if projects_found:
            result_message = "Найденные проекты:\n"
            for project in projects_found:
                result_message += (
                    f"Название: {project[1]}\n"
                    f"Преподаватель: {project[0]}\n"
                    f"Статус: {project[5]}\n"
                    f"Прогресс: {project[6]}\n"
                    f"Оценка: {project[7]}\n\n"
                )
            bot.send_message(user_id, result_message)
        else:
            bot.send_message(user_id, "Проектов с таким статусом не найдено.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при поиске проектов: {e}")

evaluation_map = {
    "Отлично":  "90EE90",
    "Хорошо": "00BFFF",
    "Удовлетворительно": "FFA500",
    "Неуд": "FF0000"
}

def get_evaluation_color(evaluation):
    return evaluation_map.get(evaluation, "FFFFFF")

def colorize_evaluations(ws):
    for row in ws.iter_rows(min_row=2):
        evaluation_cell = row[7]
        if evaluation_cell.value:
            color_code = get_evaluation_color(evaluation_cell.value)
            fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
            evaluation_cell.fill = fill

@bot.callback_query_handler(func=lambda call: call.data in ["interim_report", "final_report"])
def generate_report(call):
    user_id = call.from_user.id
    report_type = call.data

    try:
        wb = openpyxl.load_workbook(PROJECTS_FILE)
        ws = wb.active

        teacher_name = get_teacher_name(user_id)
        report_wb = openpyxl.Workbook()
        report_ws = report_wb.active
        report_ws.title = "Отчёт"

        headers = [cell.value for cell in ws[1]]
        report_ws.append(headers)

        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[0].value == teacher_name:
                if report_type == "interim_report" or (report_type == "final_report" and row[5].value == "Завершен"):
                    report_row = [cell.value for cell in row]
                    report_ws.append(report_row)

        colorize_evaluations(report_ws)
        report_file = f"{report_type}_report.xlsx"
        report_wb.save(report_file)

        with open(report_file, 'rb') as f:
            bot.send_document(user_id, f)
        os.remove(report_file)

    except Exception as e:
        bot.send_message(user_id, f"Ошибка при создании отчёта: {e}")

@bot.callback_query_handler(func=lambda call: call.data == "contact_teacher")
def contact_teacher(call):
    user_id = call.from_user.id

    if not is_student(user_id):
        bot.send_message(user_id, "Ошибка! Вы не студент.")
        return

    markup = types.InlineKeyboardMarkup()
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            teacher_id = row[0]
            teacher_name = f"{row[1]} {row[2]}"
            markup.add(types.InlineKeyboardButton(teacher_name, callback_data=f"msg_teacher_{teacher_id}"))

        bot.send_message(user_id, "Выберите преподавателя, с которым хотите связаться:", reply_markup=markup)
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при загрузке преподавателей: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("msg_teacher_"))
def teacher_selected(call):
    teacher_id = call.data.split("_")[2]
    user_id = call.from_user.id

    bot.send_message(user_id, "Введите ваше сообщение для преподавателя:")
    bot.register_next_step_handler(call.message, send_message_to_teacher, teacher_id)


def send_message_to_teacher(message, teacher_id):
    user_id = message.from_user.id
    user_message = message.text.strip()

    student_data = get_student_data(user_id)
    if student_data:
        student_lastname, student_firstname, student_group = student_data
    else:
        student_lastname = message.from_user.last_name or "Неизвестно"
        student_firstname = message.from_user.first_name or "Неизвестно"
        student_group = "Неизвестно"

    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        teacher_found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == teacher_id:
                teacher_found = True
                teacher_name = f"{row[1]} {row[2]}"
                break

        if teacher_found:
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("Ответить студенту", callback_data=f"reply_{user_id}"))
            bot.send_message(
                teacher_id,
                f"Сообщение от студента {student_lastname} {student_firstname}, группа {student_group}:\n{user_message}",
                reply_markup=markup
            )
            bot.send_message(user_id, f"Ваше сообщение было отправлено преподавателю {teacher_name}.")
        else:
            bot.send_message(user_id, "Преподаватель не найден.")
    except Exception as e:
        bot.send_message(user_id, f"Ошибка при отправке сообщения: {e}")


@bot.callback_query_handler(func=lambda call: call.data.startswith("reply_"))
def reply_to_student(call):
    teacher_id = call.from_user.id
    student_id = call.data.split("_")[1]

    if not is_teacher(teacher_id):
        bot.send_message(teacher_id, "Ошибка! Вы не преподаватель.")
        return

    bot.send_message(teacher_id, "Введите ваше сообщение для студента:")
    bot.register_next_step_handler(call.message, send_reply_to_student, student_id)


def send_reply_to_student(message, student_id):
    teacher_id = message.from_user.id
    teacher_message = message.text.strip()

    teacher_data = get_teacher_data(teacher_id)
    if teacher_data:
        teacher_lastname, teacher_firstname = teacher_data
    else:
        teacher_lastname = teacher_firstname = "Неизвестно"

    try:
        bot.send_message(
            student_id,
            f"Ответ от преподавателя {teacher_lastname} {teacher_firstname}:\n{teacher_message}"
        )
        bot.send_message(teacher_id, "Ваше сообщение было отправлено студенту.")
    except Exception as e:
        bot.send_message(teacher_id, f"Ошибка при отправке сообщения студенту: {e}")

def get_teacher_data(teacher_id):
    try:
        wb = openpyxl.load_workbook(TEACHER_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(teacher_id):
                return row[1], row[2]
        return None
    except Exception as e:
        print(f"Ошибка при получении данных преподавателя: {e}")
        return None

if __name__ == "__main__":
    bot.polling(none_stop=True)